import pytest
from selenium import webdriver
from openpyxl import load_workbook
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from LoginPage import LoginPage
from datetime import datetime

# Constants
EXCEL_FILE = 'D:\\Guvi - Automation testing\\Python folder\\O_HRM_XLSX\\login_data.xlsx'

URL = 'https://opensource-demo.orangehrmlive.com/web/index.php/auth/login'
TESTER_NAME = 'David Aravindhraj'

# Function to update Excel file with test results
def update_excel(test_id, result):
    try:
        # Load the workbook and select the active sheet
        wb = load_workbook(EXCEL_FILE)
        sheet = wb.active
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Find the row with the matching test_id and update result
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if row[0].value == test_id:
                print(f"Updating Excel for Test ID: {test_id}")  # Log message
                row[4].value = current_time.split(' ')[0]  # Date
                row[5].value = current_time.split(' ')[1]  # Time
                row[6].value = TESTER_NAME  # Tester name
                row[7].value = result  # Test result (Passed/Failed)
                break

        # Save the updated Excel file
        wb.save(EXCEL_FILE)
        print(f"Excel updated successfully for Test ID: {test_id}")

    except Exception as e:
        print(f"Error updating Excel: {e}")  # Error message

# Test case
@pytest.mark.parametrize("test_id, username, password", [
    ("T001", "Admin", "admin123"),
    ("T002", "InvalidUser1", "InvalidPass1"),
    
])
def test_login(test_id, username, password):
    driver = webdriver.Chrome()  
    driver.get(URL)
    login_page = LoginPage(driver)

    login_page.enter_username(username)
    login_page.enter_password(password)
    login_page.click_login()

    try:
        WebDriverWait(driver, 10).until(
            EC.url_contains("/dashboard")
        )
        update_excel(test_id, "Passed")
    except Exception as e:
        update_excel(test_id, "Failed")
    finally:
        driver.quit()

# Run pytest to execute the test cases
if __name__ == '__main__':
    pytest.main()
